import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from config.constants import LOCAL_FILE

def style_excel(filename):
    wb = load_workbook(filename)
    ws = wb.active

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    currency_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.fill = currency_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif isinstance(cell.value, str):
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(filename)

def process_budget(expense_selected, allotment, realignment, obligations, earmarked):
    if not os.path.exists(LOCAL_FILE):
        raise FileNotFoundError(f"{LOCAL_FILE} not found.")

    df = pd.read_excel(LOCAL_FILE)
    matched = False

    for idx, row in df.iterrows():
        if str(row.get("Expense", "")).strip() == expense_selected:
            matched = True
            df.at[idx, "Allotment"] = allotment
            df.at[idx, "Realignment"] = realignment
            df.at[idx, "Obligations"] = obligations
            df.at[idx, "Earmarked"] = earmarked

            balance = allotment + realignment + obligations
            balance_after = balance - earmarked
            utilization = obligations / allotment if allotment != 0 else 0

            df.at[idx, "Balance as of Date"] = balance
            df.at[idx, "Balance After Earmark"] = balance_after
            df.at[idx, "Utilization (%)"] = round(utilization * 100, 2)
            break

    if not matched:
        raise ValueError(f"Expense '{expense_selected}' not found.")

    df.to_excel(LOCAL_FILE, index=False)

    wb = load_workbook(LOCAL_FILE)
    ws = wb.active
    header = {cell.value: cell.column for cell in ws[1]}
    row_number = idx + 2

    if all(col in header for col in ["Allotment", "Realignment", "Obligations", "Balance as of Date"]):
        c_allot = get_column_letter(header["Allotment"])
        c_real = get_column_letter(header["Realignment"])
        c_obl = get_column_letter(header["Obligations"])
        ws.cell(row=row_number, column=header["Balance as of Date"]).value = \
            f"=ROUND({c_allot}{row_number}+{c_real}{row_number}+{c_obl}{row_number},2)"

    if all(col in header for col in ["Balance as of Date", "Earmarked", "Balance After Earmark"]):
        c_balance = get_column_letter(header["Balance as of Date"])
        c_earm = get_column_letter(header["Earmarked"])
        ws.cell(row=row_number, column=header["Balance After Earmark"]).value = \
            f"=ROUND({c_balance}{row_number}-{c_earm}{row_number},2)"

    if all(col in header for col in ["Obligations", "Allotment", "Utilization (%)"]):
        c_obl = get_column_letter(header["Obligations"])
        c_allot = get_column_letter(header["Allotment"])
        ws.cell(row=row_number, column=header["Utilization (%)"]).value = \
            f"=IF({c_allot}{row_number}=0,0,ROUND({c_obl}{row_number}/{c_allot}{row_number}*100,2))"

    wb.save(LOCAL_FILE)
    style_excel(LOCAL_FILE)
